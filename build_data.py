#!/usr/bin/env python3
"""Read wiseguys_tracker.xlsx (Best Bets + Futures tabs) and emit data.json.

The spreadsheet is the source of truth. Edit + grade there, then run:
    python3 build_data.py
...and redeploy. The dashboard reads only the generated data.json at runtime.
"""
import json, datetime, sys
from pathlib import Path
import openpyxl

SRC = Path(__file__).parent / "wiseguys_tracker.xlsx"
OUT = Path(__file__).parent / "data.json"


def clean(v):
    return "" if v is None else str(v).strip()


def payout(odds):
    """Total return on a 1u win (e.g. -110 -> 1.91, +150 -> 2.50)."""
    o = int(str(odds).replace("+", "").replace(" ", ""))
    return round(1 + (o / 100 if o > 0 else 100 / abs(o)), 2)


def net_units(odds, result):
    if result == "P": return 0.0
    if result == "L": return -1.0
    if result == "W": return round(payout(odds) - 1, 2)
    return None  # ungraded


def side_of(sel):
    s = sel.lower()
    if s.startswith("over"): return "over"
    if s.startswith("under"): return "under"
    if s == "pass": return "pass"
    return ""


# ----------------------------------------------------------------- Best Bets
def build_best_bets(wb):
    ws = wb["Best Bets"]
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}
    for n in ["Date", "Host", "Bet", "Type", "Odds", "Result", "League"]:
        if n not in idx:
            sys.exit(f"Missing column '{n}' in Best Bets tab")
    bets, dates = [], []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[idx["Date"]]:
            continue
        res = clean(row[idx["Result"]]).upper()
        odds = clean(row[idx["Odds"]])
        if res in ("W", "L", "P"):
            result, net = res, net_units(odds, res)
        elif res == "PENDING":
            result, net = "Pending", None
        else:
            continue
        bets.append({
            "date": clean(row[idx["Date"]]), "host": clean(row[idx["Host"]]),
            "type": clean(row[idx["Type"]]), "sport": clean(row[idx["League"]]),
            "bet": clean(row[idx["Bet"]]), "odds": odds, "result": result, "net": net,
        })
        dates.append(clean(row[idx["Date"]]))
    return bets, dates


# ------------------------------------------------------------------- Futures
def build_futures(wb):
    ws = wb["Futures"]
    headers = [c.value for c in ws[1]]
    ix = {h: i for i, h in enumerate(headers)}
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not clean(row[ix["Category"]]):
            continue
        rows.append({h: clean(row[ix[h]]) for h in headers})

    sports = {}
    for r in rows:
        sp = sports.setdefault(r["Sport"], {
            "win_totals": {}, "division_winners": {}, "player_props": {},
            "playoff_props": [], "longshots": [], "_rows": []})
        sp["_rows"].append(r)
        cat, div = r["Category"], r["Division"]
        odds = {"bet": r["Bet Odds"], "live": r["Live Odds"] or None}
        if cat == "Win Total":
            board = sp["win_totals"].setdefault(div, {})
            team = board.setdefault(r["Team"], {"team": r["Team"], "line": r["Line"], "picks": []})
            team["picks"].append({"host": r["Host"], "side": side_of(r["Selection"]),
                "selection": r["Selection"], "odds": odds, "stake": r["Stake"],
                "alt": r["Bet Type"] == "Alternate Win Total"})
        elif cat == "Division Winner":
            sp["division_winners"].setdefault(div, []).append(
                {"host": r["Host"], "team": r["Team"], "odds": odds, "result": r["Result"]})
        elif cat == "Player Prop":
            sp["player_props"].setdefault(div, []).append(
                {"host": r["Host"], "bet": r["Selection"], "odds": odds, "result": r["Result"]})
        elif cat == "Playoff Prop":
            sp["playoff_props"].append({"host": r["Host"], "bet": r["Selection"],
                "bet_type": r["Bet Type"], "division": div, "odds": odds, "result": r["Result"]})
        elif cat == "Longshot":
            sp["longshots"].append({"host": r["Host"], "bet": r["Selection"],
                "bet_type": r["Bet Type"], "division": div, "odds": odds, "result": r["Result"]})

    for sp in sports.values():
        # consensus per win-total team
        for div, teams in sp["win_totals"].items():
            out = []
            for t in teams.values():
                picks = [p for p in t["picks"] if p["side"] != "pass"]
                sides = {p["side"] for p in picks}
                t["consensus"] = ("family" if len(sides) == 1 else "split") if len(picks) >= 2 else ""
                out.append(t)
            sp["win_totals"][div] = out
        # per-sport summary
        rws = sp.pop("_rows")
        plays = [r for r in rws if r["Stake"] == "Play"]
        market = ("Division Winner", "Player Prop", "Playoff Prop")  # non-longshot staked
        sp["summary"] = {
            "open_predictions": len({(r["Division"], r["Team"], r["Category"], r["Selection"]) for r in rws}),
            "staked_plays": len(plays),
            "longshots": len([r for r in rws if r["Category"] == "Longshot"]),
            "max_units": round(sum(payout(r["Bet Odds"]) for r in rws
                                   if r["Category"] in market and r["Bet Odds"]), 1),
        }
    return sports


def main():
    wb = openpyxl.load_workbook(SRC, data_only=False)
    bets, dates = build_best_bets(wb)
    futures = build_futures(wb)
    data = {
        "meta": {
            "generated": datetime.datetime.now().isoformat(timespec="minutes"),
            "range": f"{dates[0]}\u2013{dates[-1]}" if dates else "",
            "count": len(bets), "source": "wiseguys_tracker.xlsx",
        },
        "bets": bets,
        "futures": {"sports": futures},
    }
    OUT.write_text(json.dumps(data, indent=2))
    fp = {s: d["summary"] for s, d in futures.items()}
    print(f"Wrote {OUT.name}: {len(bets)} best bets; futures per sport:")
    for s, sm in fp.items():
        print(f"  {s}: {sm['open_predictions']} open, {sm['staked_plays']} staked, "
              f"{sm['longshots']} longshots, max {sm['max_units']}u")


if __name__ == "__main__":
    main()
